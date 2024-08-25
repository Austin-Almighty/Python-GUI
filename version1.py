import re
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, Button, Label, filedialog, messagebox

def check_date_format(date_series):
    pattern = re.compile(r'^\d{8}$')
    def is_valid_date(date_str):
        if not pattern.match(date_str):
            return False
        try:
            datetime.strptime(date_str, '%Y%m%d')
            return True
        except ValueError:
            return False
    return date_series.apply(lambda x: is_valid_date(str(x)))

def check_scores(score, max_value):
    return 0 <= score <= max_value and score % 5 == 0

def check_pass(*scores):
    return 'pass' if all(score >= 70 for score in scores) else 'did not pass'

def process_file(file_path):
    df = pd.read_excel(file_path, sheet_name='Sheet1')

    # Validate date of birth and test date
    df['DOB_Format_Correct'] = check_date_format(df['date_of_birth'])
    df['Test_Date_Format_Correct'] = check_date_format(df['test_date'])

    # Validate scores
    score_columns = [
        ('abacus_multiplication', 100),
        ('abacus_division', 100),
        ('abacus_addition_subtraction', 150),
        ('mental_multiplication', 100),
        ('mental_division', 100),
        ('mental_addition_subtraction', 150)
    ]

    for col, max_val in score_columns:
        df[col + '_Valid'] = df[col].apply(lambda x: check_scores(x, max_val))

    # Determine pass/fail status
    df['abacus_pass'] = df.apply(lambda row: check_pass(row['abacus_multiplication'], row['abacus_division'], row['abacus_addition_subtraction']), axis=1)
    df['mental_pass'] = df.apply(lambda row: check_pass(row['mental_multiplication'], row['mental_division'], row['mental_addition_subtraction']), axis=1)

    # Save the DataFrame back to Excel to maintain formatting and then apply styles
    file_dir, file_name = os.path.split(file_path)
    file_base, file_ext = os.path.splitext(file_name)
    output_file_path = os.path.join(file_dir, f"{file_base}_verified{file_ext}")
    df.to_excel(output_file_path, index=False, sheet_name='Sheet1')

    # Load the workbook and apply styles
    wb = load_workbook(output_file_path)
    ws = wb['Sheet1']
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    dob_column_index = df.columns.get_loc('date_of_birth') + 1
    test_date_column_index = df.columns.get_loc('test_date') + 1

    for row in range(2, ws.max_row + 1):
        dob_cell = ws.cell(row=row, column=dob_column_index)
        test_date_cell = ws.cell(row=row, column=test_date_column_index)
        if not df.at[row - 2, 'DOB_Format_Correct']:
            dob_cell.fill = red_fill
        if not df.at[row - 2, 'Test_Date_Format_Correct']:
            test_date_cell.fill = red_fill

    for col, max_val in score_columns:
        col_index = df.columns.get_loc(col) + 1
        for row in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row, column=col_index)
            if not df.at[row - 2, col + '_Valid']:
                score_cell.fill = red_fill

    abacus_pass_index = df.columns.get_loc('abacus_pass') + 1
    mental_pass_index = df.columns.get_loc('mental_pass') + 1

    for row in range(2, ws.max_row + 1):
        abacus_cell = ws.cell(row=row, column=abacus_pass_index)
        mental_cell = ws.cell(row=row, column=mental_pass_index)
        if df.at[row - 2, 'abacus_pass'] == 'did not pass':
            abacus_cell.fill = red_fill
        if df.at[row - 2, 'mental_pass'] == 'did not pass':
            mental_cell.fill = red_fill

    wb.save(output_file_path)
    messagebox.showinfo("Success", f"File processed and saved as {output_file_path}")

def open_file():
    file_path = filedialog.askopenfilename(title="Open Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        process_file(file_path)

# Create the main window
root = Tk()
root.title("Excel Verification Tool")

# Create a button to open the file
btn_open = Button(root, text="Open Excel File", command=open_file)
btn_open.pack(pady=20)

# Create a label to show the status
label = Label(root, text="Please open an Excel file to start verification.")
label.pack(pady=10)

# Run the application
root.mainloop()