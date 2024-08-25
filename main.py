import re
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = '/mnt/data/testing.xlsx'
df = pd.read_excel(file_path, sheet_name='Sheet1')

# Check for Birth Date and Test Date Format

def check_date_format(date_series):
    """
    Check if the date in the series follows the YYYYMMDD format and is a valid date.
    
    Args:
    date_series (pd.Series): Series containing the date.

    Returns:
    pd.Series: Series indicating whether the format and date are correct.
    """
    pattern = re.compile(r'^\d{8}$')
    
    def is_valid_date(date_str):
        if not pattern.match(date_str):
            return False
        try:
            datetime.strptime(date_str, '%Y%m%d')
            return True
        except ValueError:
            return False
    
    return date_series.apply(lambda x: is_valid_date(str(x)))

# Apply the function to the 'date_of_birth' and 'test_date' columns
df['DOB_Format_Correct'] = check_date_format(df['date_of_birth'])
df['Test_Date_Format_Correct'] = check_date_format(df['test_date'])

# Check Scores

def check_scores(score, max_value):
    """
    Check if the score is a multiple of 5 and within the allowed range.

    Args:
    score (int): The score to be checked.
    max_value (int): The maximum valid score.

    Returns:
    bool: True if the score is valid, False otherwise.
    """
    return 0 <= score <= max_value and score % 5 == 0

# Apply the score checking function
score_columns = [
    ('abacus_multiplication', 100),
    ('abacus_division', 100),
    ('abacus_addition_subtraction', 150),
    ('mental_multiplication', 100),
    ('mental_division', 100),
    ('mental_addition_subtraction', 150)
]

for col, max_val in score_columns:
    df[col + '_Valid'] = df[col].apply(lambda x: check_scores(x, max_val))

# Determine if the student passes the Abacus and Mental Arithmetic components

def check_pass(*scores):
    """
    Check if a student passes by getting at least 70 in all three subjects.

    Args:
    scores: Variable length argument list of scores.

    Returns:
    str: 'pass' if all scores are >= 70, 'did not pass' otherwise.
    """
    return 'pass' if all(score >= 70 for score in scores) else 'did not pass'

df['abacus_pass'] = df.apply(lambda row: check_pass(row['abacus_multiplication'], row['abacus_division'], row['abacus_addition_subtraction']), axis=1)
df['mental_pass'] = df.apply(lambda row: check_pass(row['mental_multiplication'], row['mental_division'], row['mental_addition_subtraction']), axis=1)

# Save the DataFrame back to Excel to maintain formatting and then apply styles
output_file_path = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}_error.xlsx")
df.to_excel(output_file_path, index=False, sheet_name='Sheet1')

# Load the workbook and select the sheet
wb = load_workbook(output_file_path)
ws = wb['Sheet1']

# Define the fill style for incorrect dates and invalid scores
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Apply the fill style to cells with incorrect dates
dob_column_index = df.columns.get_loc('date_of_birth') + 1
test_date_column_index = df.columns.get_loc('test_date') + 1

for row in range(2, ws.max_row + 1):
    dob_cell = ws.cell(row=row, column=dob_column_index)
    test_date_cell = ws.cell(row=row, column=test_date_column_index)
    if not df.at[row - 2, 'DOB_Format_Correct']:
        dob_cell.fill = red_fill
    if not df.at[row - 2, 'Test_Date_Format_Correct']:
        test_date_cell.fill = red_fill

# Apply the fill style to cells with invalid scores
for col, max_val in score_columns:
    col_index = df.columns.get_loc(col) + 1
    for row in range(2, ws.max_row + 1):
        score_cell = ws.cell(row=row, column=col_index)
        if not df.at[row - 2, col + '_Valid']:
            score_cell.fill = red_fill

# Apply the fill style to the "did not pass" status
abacus_pass_index = df.columns.get_loc('abacus_pass') + 1
mental_pass_index = df.columns.get_loc('mental_pass') + 1

for row in range(2, ws.max_row + 1):
    abacus_cell = ws.cell(row=row, column=abacus_pass_index)
    mental_cell = ws.cell(row=row, column=mental_pass_index)
    if df.at[row - 2, 'abacus_pass'] == 'did not pass':
        abacus_cell.fill = red_fill
    if df.at[row - 2, 'mental_pass'] == 'did not pass':
        mental_cell.fill = red_fill

# Save the workbook
wb.save(output_file_path)